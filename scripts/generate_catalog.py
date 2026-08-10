#!/usr/bin/env python3
"""
Azam Reka — Catalogue Generator
Reads azam-reka-products.xlsx and regenerates the product grid inside catalog.html.
Supports up to 4 images per product; auto-adds carousel arrows/dots when a product
has more than one image.

Usage:
    python3 generate_catalog.py path/to/azam-reka-products.xlsx path/to/catalog.html
"""

import sys
import re
import html
import os
import json
import openpyxl


def esc(text):
    """HTML-escape a string, safely handling None."""
    if text is None:
        return ""
    return html.escape(str(text))


def slugify(name):
    """Turn a product name into a URL-safe filename slug."""
    slug = re.sub(r"[^a-z0-9]+", "-", str(name).lower()).strip("-")
    return slug or "product"


def is_video_file(path):
    """Check if a media path looks like a video based on extension."""
    if not path:
        return False
    video_extensions = (".mp4", ".webm", ".mov", ".m4v", ".ogg")
    return str(path).lower().strip().endswith(video_extensions)


def build_media_html(images, video, alt_text, on_sale=False):
    """Build the product-card__media block — single image, or carousel with optional video slide."""
    images = [img for img in images if img]  # drop blanks
    if not images:
        images = ["https://images.unsplash.com/photo-1595079676339-1534801ad6cf?w=700&q=80"]

    media_items = [{"type": "image", "src": img} for img in images]
    if video:
        media_items.append({"type": "video", "src": video})

    sale_badge = '<span class="sale-badge">Sale</span>\n          ' if on_sale else ""

    if len(media_items) == 1:
        item = media_items[0]
        if item["type"] == "video":
            return f'''<div class="product-card__media kerf-card">
          {sale_badge}<video src="{esc(item['src'])}" muted loop playsinline preload="metadata" controls></video>
        </div>'''
        return f'''<div class="product-card__media kerf-card">
          {sale_badge}<img src="{esc(item['src'])}" alt="{esc(alt_text)}" loading="lazy">
        </div>'''

    slides_parts = []
    for i, item in enumerate(media_items):
        active_class = " is-active" if i == 0 else ""
        if item["type"] == "video":
            slides_parts.append(f'''            <div class="carousel-slide{active_class}">
              <span class="video-badge">Video</span>
              <video src="{esc(item['src'])}" muted loop playsinline preload="metadata" controls></video>
            </div>''')
        else:
            slides_parts.append(f'''            <div class="carousel-slide{active_class}">
              <img src="{esc(item['src'])}" alt="{esc(alt_text)} photo {i+1}" loading="lazy">
            </div>''')
    slides = "\n".join(slides_parts)

    dots = "\n".join(
        f'            <button class="carousel-dot{" is-active" if i == 0 else ""}" data-index="{i}" aria-label="View item {i+1}"></button>'
        for i in range(len(media_items))
    )

    return f'''<div class="product-card__media kerf-card" data-carousel>
          {sale_badge}<div class="carousel-track">
{slides}
          </div>
          <button class="carousel-arrow carousel-arrow--prev" aria-label="Previous">&#8249;</button>
          <button class="carousel-arrow carousel-arrow--next" aria-label="Next">&#8250;</button>
          <div class="carousel-dots">
{dots}
          </div>
        </div>'''


def build_product_card(row):
    row = row[:13]
    (sku, name, category, price, sale_price, tag,
     img1, img2, img3, img4, video, description, active) = row

    if not name or not category or price is None:
        return None
    if active is not None and str(active).strip().upper() == "N":
        return None

    on_sale = bool(sale_price)
    images = [img1, img2, img3, img4]
    media_html = build_media_html(images, video, name, on_sale=on_sale)

    tag_html = f'<span class="product-card__tag">{esc(tag)}</span>\n        ' if tag else ""

    if sale_price:
        price_html = (f'RM {float(price):.2f} '
                      f'<span style="text-decoration: line-through; color: var(--color-structural); '
                      f'margin-left: 0.4em;">RM {float(sale_price):.2f}</span>')
        cart_price = float(price)
    else:
        price_html = f'RM {float(price):.2f}'
        cart_price = float(price)

    return f'''      <div class="product-card" data-category="{esc(category)}">
        <a href="products/{slugify(name)}.html" style="display:block;">
          {media_html}
        </a>
        {tag_html}<a href="products/{slugify(name)}.html" style="text-decoration:none; color:inherit;"><span class="product-card__title">{esc(name)}</span></a>
        <span class="mono-price">{price_html}</span>
        <button data-add-to-cart="{esc(name)}" data-price="{cart_price:.2f}" class="btn btn-whatsapp" style="margin-top: 0.5rem;">Add to Cart</button>
      </div>'''


PRODUCT_PAGE_TEMPLATE = '''<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{name} — Azam Reka</title>
<meta name="description" content="{description}">
<link rel="icon" href="../assets/logo.png">
<link rel="stylesheet" href="../assets/styles.css">
</head>
<body>

<div class="promo-bar">
  <button class="promo-arrow promo-arrow--prev" aria-label="Previous announcement">&#8249;</button>
  <div class="promo-track">
    <div class="promo-slide is-active">Custom Wedding &amp; Hari Guru orders now open &mdash; <a href="/contact.html">WhatsApp for a quote</a></div>
    <div class="promo-slide">Rated 5.0 on Google &mdash; loved by teachers &amp; newlyweds</div>
    <div class="promo-slide">Placed an order? <a href="/track-order.html">Track it here</a></div>
  </div>
  <button class="promo-arrow promo-arrow--next" aria-label="Next announcement">&#8250;</button>
</div>

<header class="site-header" data-scrolled="false">
  <div class="container site-header__inner">
    <a href="../index.html" class="site-header__logo" aria-label="Azam Reka">
      <img src="../assets/logo.png" alt="Azam Reka" style="height:42px; width:auto;">
    </a>
    <nav class="site-header__nav">
      <a href="../catalog.html">Catalogue</a>
      <a href="../about.html">About</a>
      <a href="../contact.html">Custom Orders</a>
    </nav>
    <div class="site-header__actions" style="display:flex; align-items:center; gap: 1rem;">
      <button data-cart-btn class="cart-button" aria-label="Shopping cart" style="background:none; border:none; cursor:pointer; font-size:0;">
        <svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
          <circle cx="9" cy="21" r="1"/><circle cx="20" cy="21" r="1"/>
          <path d="M1 1h4l2.68 13.39a2 2 0 0 0 2 1.61h9.72a2 2 0 0 0 2-1.61L23 6H6"/>
        </svg>
        <span data-cart-count class="cart-count"></span>
      </button>
      <a href="https://wa.me/601110852324" class="btn btn-primary" style="padding: 0.6em 1.2em;">WhatsApp Us</a>
      <button class="site-header__mobile-toggle" data-mobile-nav-toggle aria-label="Open menu" aria-expanded="false">&#9776;</button>
    </div>
  </div>
  <div class="mobile-nav" data-mobile-nav data-open="false" style="display:none; flex-direction:column; padding: 1rem clamp(1.25rem,4vw,4rem); border-top: 1px solid var(--color-ash); background: var(--color-bone);">
    <a href="../catalog.html" style="padding-block:0.5em;">Catalogue</a>
    <a href="../about.html" style="padding-block:0.5em;">About</a>
    <a href="../contact.html" style="padding-block:0.5em;">Custom Orders</a>
  </div>
</header>

<section class="section">
  <div class="container">
    <p class="body-sm" style="color: var(--color-structural); margin-bottom: var(--space-3);">
      <a href="../catalog.html" style="color: var(--color-structural);">Catalogue</a> &rsaquo; {name}
    </p>
    <div class="grid-2" style="align-items: flex-start; gap: var(--space-5);">
      <div>
        {media_html}
      </div>
      <div>
        {tag_html}
        <h1 class="h-display-3" style="margin-top: 0.4em;">{name}</h1>
        <p class="mono-price" style="font-size: 1.3rem; margin-top: 0.6em;">{price_html}</p>
        <p class="body-base" style="margin-top: var(--space-3); max-width: 46ch; color: var(--color-char);">{description}</p>
        <div style="display:flex; flex-direction: column; gap: 0.75rem; margin-top: var(--space-4); max-width: 340px;">
          <button data-add-to-cart="{name_attr}" data-price="{cart_price:.2f}" class="btn btn-whatsapp">Add to Cart</button>
          <a href="https://wa.me/601110852324?text=Hi%20Azam%20Reka%2C%20I%27d%20like%20to%20ask%20about%3A%20{name_url}" class="btn btn-outline">Ask a Question on WhatsApp</a>
        </div>
      </div>
    </div>
  </div>
</section>

<!-- CART MODAL -->
<div data-cart-modal class="cart-modal" style="display:none;">
  <div class="cart-modal__content">
    <button data-cart-close class="cart-modal__close">×</button>
    <h2 class="h-heading" style="margin-top:0;">Your Cart</h2>
    <div data-cart-items style="display:none;"></div>
    <div data-cart-empty style="text-align:center; padding: 2rem 0;">
      <p class="body-base" style="color: var(--color-structural);">Your cart is empty.</p>
    </div>
    <div data-cart-checkout style="display:none; flex-direction:column; gap: 1rem; margin-top: 2rem; padding-top: 1.5rem; border-top: 1px solid var(--color-ash);">
      <div style="display:flex; justify-content:space-between; align-items:center;">
        <p class="h-heading" style="margin:0;">Total</p>
        <p class="h-heading" style="margin:0;" data-cart-total>RM 0.00</p>
      </div>
      <button data-cart-whatsapp class="btn btn-primary" style="width:100%;">Proceed to Checkout</button>
    </div>
  </div>
</div>

<footer class="site-footer">
  <div class="container">
    <div class="site-footer__grid">
      <div>
        <div class="h-heading" style="color: var(--color-bone); margin-bottom: var(--space-2);">AZAM REKA</div>
        <p class="body-sm" style="color: var(--color-structural); max-width: 34ch;">
          Custom laser-cutting and engraving, crafted with passion, precision, and a touch of artistry.
        </p>
      </div>
      <div>
        <p class="mono-label" style="color: var(--color-bone); margin-bottom: var(--space-2);">Catalogue</p>
        <a href="../catalog.html?category=plaques">Plaques &amp; Awards</a>
        <a href="../catalog.html?category=keychains">Keychains &amp; Accessories</a>
        <a href="../catalog.html?category=wedding">Wedding &amp; Mas Kahwin</a>
        <a href="../catalog.html?category=decor">Home D&eacute;cor</a>
      </div>
      <div>
        <p class="mono-label" style="color: var(--color-bone); margin-bottom: var(--space-2);">Reach Us</p>
        <a href="https://wa.me/601110852324">WhatsApp</a>
        <a href="https://instagram.com/azam.reka">Instagram @azam.reka</a>
        <a href="mailto:azam.r3ka@gmail.com">Email</a>
      </div>
    </div>
    <div class="site-footer__bottom">
      <span>&copy; 2026 Azam Reka. All rights reserved.</span>
    </div>
  </div>
</footer>

<a href="https://wa.me/601110852324" class="whatsapp-float" aria-label="Chat on WhatsApp">
  <svg width="28" height="28" viewBox="0 0 24 24" fill="currentColor"><path d="M17.5 14.4c-.3-.1-1.7-.9-2-1-.3-.1-.5-.1-.7.1-.2.3-.8 1-.9 1.1-.2.2-.3.2-.6.1-.3-.1-1.2-.5-2.3-1.5-.9-.8-1.4-1.7-1.6-2-.2-.3 0-.5.1-.6.1-.1.3-.3.4-.5.1-.1.2-.3.3-.4.1-.2 0-.4 0-.5C10.1 9 9.6 7.7 9.4 7.2c-.2-.5-.4-.4-.5-.4h-.5c-.2 0-.5.1-.7.3-.3.3-1 1-1 2.4s1 2.8 1.2 3c.1.2 2 3 4.8 4.3.7.3 1.2.5 1.6.6.7.2 1.3.2 1.8.1.5-.1 1.7-.7 1.9-1.3.2-.7.2-1.2.2-1.3-.1-.2-.3-.2-.6-.4zM12 2C6.5 2 2 6.5 2 12c0 1.9.5 3.6 1.4 5.2L2 22l4.9-1.3C8.4 21.5 10.1 22 12 22c5.5 0 10-4.5 10-10S17.5 2 12 2z"/></svg>
</a>

<script src="../assets/main.js"></script>
<script src="../assets/engraving-preview.js"></script>
</body>
</html>
'''


def build_product_detail_page(row):
    row = row[:13]
    (sku, name, category, price, sale_price, tag,
     img1, img2, img3, img4, video, description, active) = row

    if not name or not category or price is None:
        return None, None
    if active is not None and str(active).strip().upper() == "N":
        return None, None

    on_sale = bool(sale_price)
    images = [img1, img2, img3, img4]
    media_html = build_media_html(images, video, name, on_sale=on_sale)

    tag_html = f'<span class="product-card__tag"{" style=\"color: var(--color-brass); border-color: var(--color-brass-soft);\"" if category == "wedding" else ""}>{esc(tag)}</span>' if tag else ""

    if sale_price:
        price_html = (f'RM {float(price):.2f} '
                      f'<span style="text-decoration: line-through; color: var(--color-structural); '
                      f'margin-left: 0.4em; font-size: 0.85em;">RM {float(sale_price):.2f}</span>')
        cart_price = float(price)
    else:
        price_html = f'RM {float(price):.2f}'
        cart_price = float(price)

    desc_text = description if description else f"Custom laser-cut {esc(name).lower()}, made to order by Azam Reka."

    html_out = PRODUCT_PAGE_TEMPLATE.format(
        name=esc(name),
        name_attr=esc(name),
        name_url=str(name).replace(" ", "%20"),
        description=esc(desc_text),
        media_html=media_html,
        tag_html=tag_html,
        price_html=price_html,
        cart_price=cart_price,
    )
    return slugify(name), html_out


def build_zoho_item_map(ws):
    """Read the Zoho Item ID column and build a name->item_id map for the backend."""
    header_row = [cell.value for cell in ws[1]]
    try:
        name_col = header_row.index("Product Name")
        zoho_id_col = header_row.index("Zoho Item ID")
    except ValueError:
        return {}

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_col]
        zoho_id = row[zoho_id_col] if zoho_id_col < len(row) else None
        if name and zoho_id:
            mapping[str(name).strip()] = str(zoho_id).strip()
    return mapping


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 generate_catalog.py <products.xlsx> <catalog.html>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    html_path = sys.argv[2]
    site_dir = os.path.dirname(os.path.abspath(html_path))
    products_dir = os.path.join(site_dir, "products")
    os.makedirs(products_dir, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Products"]

    cards = []
    detail_pages = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        card = build_product_card(row)
        if card:
            cards.append(card)
        slug, page_html = build_product_detail_page(row)
        if slug:
            detail_pages.append((slug, page_html))

    if not cards:
        print("Warning: no active products found in spreadsheet. Aborting to avoid wiping catalog.")
        sys.exit(1)

    products_html = "\n\n".join(cards)

    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()

    pattern = re.compile(
        r"(<!-- PRODUCTS_START -->\n)(.*?)(\n<!-- PRODUCTS_END -->)",
        re.DOTALL
    )

    if not pattern.search(content):
        print("Error: PRODUCTS_START / PRODUCTS_END markers not found in catalog.html")
        sys.exit(1)

    new_content = pattern.sub(lambda m: m.group(1) + products_html + m.group(3), content)

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(new_content)

    for slug, page_html in detail_pages:
        page_path = os.path.join(products_dir, f"{slug}.html")
        with open(page_path, "w", encoding="utf-8") as f:
            f.write(page_html)

    zoho_map = build_zoho_item_map(ws)
    zoho_map_path = os.path.join(site_dir, "api", "zoho-item-map.json")
    os.makedirs(os.path.dirname(zoho_map_path), exist_ok=True)
    with open(zoho_map_path, "w", encoding="utf-8") as f:
        json.dump(zoho_map, f, indent=2)

    print(f"Generated {len(cards)} product cards into {html_path}")
    print(f"Generated {len(detail_pages)} product detail pages into {products_dir}/")
    print(f"Wrote {len(zoho_map)} Zoho item mappings to {zoho_map_path}")


if __name__ == "__main__":
    main()
